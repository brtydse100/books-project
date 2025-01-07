import re
from docx import Document
import openpyxl
import requests
import xlsxwriter
import os
from .constants import *


def get_books_names(doc_file_path):
    doc = Document(doc_file_path)
    titles = []
    books_title = []
    books_in_sheet = []
    authors = []
    books = 0
    years = []
    count_books = 0
    check = 0

    book_title = ""

    for paragraph in doc.paragraphs:
        book_title = ""
        all_bold = True
        text = paragraph.text
        for run in paragraph.runs:

            if run.font.underline:
                print("underline")
                pass

            elif run.bold:
                book_title = book_title + run.text

            else:
                all_bold = False

        if (
            all_bold and text.strip() and ">>" not in text
        ):  # checks if the line is a title or a book name
            title = text
            if check == 1 and books == 0:
                print("pop")
                titles.pop()
                titles.append(title)
                check = 0

            elif books == 0:
                check += 1
                titles.append(title)

            if books != 0 and ">>" not in title:
                titles.append(title)
                books_in_sheet.append(books)

                books = 0
                check = 0

        elif book_title:
            book_title = book_title.strip(" ")
            book_title = book_title.strip(",")
            books_title.append(book_title)
            year = re.findall(r"\d+", text)
            if year == []:
                years.append(0)
            elif isinstance(year, list):
                years.append(year[0])
            else:
                years.append(year)
            books += 1
            count_books += 1
            check = 0

            author = get_author(text, book_title)

            authors.append(author)

    if books > 0:
        books_in_sheet.append(books)

    return books_title, titles, books_in_sheet, authors, years


def get_author(paragraph, book_name):
    paragraph = paragraph.replace(book_name, "")
    words = paragraph.split(",")

    words = [word.strip(" ") for word in words]

    for word in words:
        if word != "" and ("2" not in word and "1" not in word):
            return word


def book_name_normelized(book_title):

    new_book_title = ""
    for letter in book_title:
        if ord(letter) < 1400:
            letter = hex(ord(letter)).replace("0x", "%")
            new_book_title = new_book_title + letter

        else:
            new_book_title = new_book_title + letter

    return new_book_title


def get_hebrew_date(date):
    year = 1240
    date = date.replace('"', "")
    date = date.replace("\\", "")

    for letter in date:
        year = year + HEBREW_GEMATRIA[letter]

    return year


def get_date(date):
    year = re.findall(r"\d+", date)

    if year:
        year = year[0]
        return year

    text = date.split(" ")

    for year in text:
        if '"' in year:
            return get_hebrew_date((str)(year))

    return False


def find_book_by_year(found_date=0, year=0):
    found_date = (str)(get_date(found_date))
    year = (str)(year)

    if not found_date:
        return False

    if year != 0 and found_date != 0:
        if year not in found_date:
            return False

    return True


def find_book_by_author(author, book_title, found_title):

    book_name_found = found_title.split("/")[0]

    new_author = author.replace("-", " ")
    new_author = new_author.split(" ")
    # print(new_author)
    # print(author)
    for name in new_author:
        if name not in found_title:
            return False

    if book_title in book_name_found:
        return True
    return True


def find_book(author, found_title, found_date=0, year=0):
    if find_book_by_year(found_date, year) and find_book_by_author(author, found_title):
        return True

    if find_book_by_year(found_date, year):
        return True

    return False

def check_api_reponse(base_url):
    response = requests.get(base_url)
    if response.status_code == 200 and response.json() != []:
        books_data = response.json()
        return books_data
    
    return False
        
    
def api_response(base_url, book_title = None, author = None, year = 0):
    books_data = check_api_reponse(base_url)
    new_books_data = []
        
    if books_data:

        if len(books_data) == 1:
            book_data = data_organizer(books_data)
            return True, book_data

        if len(books_data) > 1:

            for book in books_data:
                book_data = data_organizer([book])
                try:
                    # found_book = find_book(author, book_data["title"], book_data["date"], year)
                    found_book = find_book_by_year(book_data["date"], year)
                except Exception as e:
                    found_book = False

                if found_book:
                    new_books_data.append(book_data)

            if len(new_books_data) == 0:

                for book in books_data:
                    book_data = data_organizer([book])
                    if author:
                        found_book = find_book_by_author(
                            author, book_title, book_data["title"]
                        )
                    else:
                        found_book = False

                    if found_book:
                        new_books_data.append(book_data)

            if len(new_books_data) == 1:
                return True, new_books_data[0]

            if len(new_books_data) > 1:
                print("found 2 books with the same date")

                for book_data in new_books_data:
                    if author:
                        found_book = find_book_by_author(
                            author, book_title, book_data["title"]
                        )
                    else:
                        found_book = False

                    if not found_book:
                        print("found a book with the same date but wrong author")
                        new_books_data.remove(book_data)

            if len(new_books_data) == 1 or len(new_books_data) > 1:
                return True, new_books_data[0]

    return False, None


def search_book_in_database(book_title, author, year):
    changed_book_title = book_name_normelized(book_title)
    if author:
        changed_author = book_name_normelized(author)
    else:
        changed_author = ""

    if year != 0:
        first_url = f"https://api.nli.org.il/openlibrary/search?api_key={USER_KEY}&query=title,exact,{changed_book_title},AND;language,contains,heb&material_type=book"
        second_url = f"https://api.nli.org.il/openlibrary/search?api_key={USER_KEY}&query=title,contains,{changed_book_title},AND;language,contains,heb&material_type=book"
        third_url = f"https://api.nli.org.il/openlibrary/search?api_key={USER_KEY}&query=title,contains,{changed_book_title},AND;language,contains,heb,AND;creator,contains,{changed_author},AND;start_date,contains,{year}&material_type=book"
        urls = [third_url, first_url, second_url]

        for url in urls:
            # print(f"{year} ---  {book_title}")
            response, data = api_response(url, book_title, author, year)
            if response:
                return data

            # print(url)

    return False


def data_organizer(book_data):
    if not book_data or not isinstance(book_data, list) or len(book_data) == 0:
        return {}

    data = book_data[0]
    book_info = {}

    for detail, values in data.items():
        field = detail.replace("http://purl.org/dc/elements/1.1/", "")

        if field == "non_standard_date":
            field = "date"

        if field not in FIELDS:
            continue

        for value_dict in values:
            value = value_dict.get("@value") or value_dict.get("@id")

            if value:
                if field in book_info:
                    if isinstance(book_info[field], list):
                        if value not in book_info[field]:
                            book_info[field].append(value)
                    else:
                        book_info[field] = [book_info[field], value]
                else:
                    book_info[field] = value

    return book_info


def clean_sheet_title(title):

    title = title.replace("*", "")
    title = title.replace("הספרים המוצעים", "")
    title = title.replace("'", "")

    return title


def delete_file_if_exists(filepath):
    # Check if the file exists
    if os.path.exists(filepath):
        try:
            # Attempt to delete the file
            os.remove(filepath)
            print(f"File '{filepath}' has been deleted.")
        except OSError as e:
            print(f"Error deleting file '{filepath}': {e}")
    else:
        print(f"File '{filepath}' does not exist.")


def get_xlsx(docs_file_path, sheet, batch_name, row_num):
    books_title, titles, books_in_sheet, authors, years = get_books_names(
        docs_file_path
    )
    
    col_num = 0
    count = 0
    row_num = row_num

    for num, title in enumerate(titles):
        for temp in range(books_in_sheet[num]):  # Example: 20 books

            print(f"{count + 1} / {len(books_title)}")

            book_title = books_title[count]
            author = authors[count]
            year = years[count]

            book_info = search_book_in_database(book_title, author, year)
            
            if book_info:
                write_row(book_info,sheet,row_num,batch_name,title)
            else:
                write_missing_rows(book_title,author,sheet,col_num,row_num,title,batch_name)

            row_num += 1
            count += 1

    return count


def write_row(book_info, sheet,row_num, batch_name, title):
    book = []

    for field in FIELDS:
        if field in book_info:
            if isinstance(book_info[field], list):
                x = True
                for item in book_info[field]:
                    if item is not None and x:
                        book_info[field] = item
                        x = False

            book.append(book_info[field])
        else:
            book.append(" ")

    for col_num, item in enumerate(book, start=1):

        sheet.cell(row=row_num, column=col_num).value = item

    sheet.cell(row=row_num, column=col_num + 1).value = batch_name
    sheet.cell(row=row_num, column=col_num + 2).value = title


def write_missing_rows(book_title,author,sheet,col_num,row_num,title,batch_name):
    print(f"didnt find: {book_title} / {author}")
    sheet.cell(row=row_num, column=1).value = f"{book_title} / {author}"
    sheet.cell(row=row_num, column=col_num + 1).value = batch_name
    sheet.cell(row=row_num, column=col_num + 2).value = title


def write_headers(sheet):
    for col_num, field in enumerate(FIELDS, start=1):
        sheet.cell(row=1, column=col_num).value = field

        sheet.cell(row=1, column=col_num + 1).value = "book docs year"
        sheet.cell(row=1, column=col_num + 2).value = "book under title"

        
def get_all_years(xlsx_file_path, docs_files_path, name="main"):

    row_num = 2
    col_num = 0
    # delete_file_if_exists(xlsx_file_path)

    workbook = xlsxwriter.Workbook(xlsx_file_path)
    workbook.close()
    workbook = openpyxl.load_workbook(xlsx_file_path)

    if "Sheet1" in workbook.sheetnames:
        workbook.remove(workbook["Sheet1"])

    sheet_title = clean_sheet_title(name)
    sheet = workbook.create_sheet(sheet_title)

    print(f"created a new sheet: {sheet_title}")

    # creates headers for the excel sheet
    write_headers(sheet)
    
    if isinstance(docs_files_path, list):
        for docs_file_path in docs_files_path:
            batch_name = os.path.basename(docs_file_path).replace(".docx", "")
            temp = get_xlsx(docs_file_path, sheet, batch_name, row_num)
            row_num += temp
            workbook.save(xlsx_file_path)
    else:
        batch_name = os.path.basename(docs_files_path).replace(".docx", "")
        temp = get_xlsx(docs_files_path, sheet, batch_name, row_num)
        row_num += temp
        workbook.save(xlsx_file_path)


def excel_completer(xlsx_file_path):
    excel_file = openpyxl.load_workbook(xlsx_file_path)

    sheet = excel_file.active
    max_rows = sheet.max_row

    for row_num in range(1, max_rows):
        cell_value = sheet.cell(row=row_num, column=5).value
        if not cell_value:
            # print(sheet.cell(row=row_num, column=1).value)
            system_number = sheet.cell(row=row_num, column=12).value
            batch_name = sheet.cell(row=row_num, column=15).value
            title = sheet.cell(row=row_num, column=16).value
        
            url = f"https://api.nli.org.il/openlibrary/search?api_key={USER_KEY}&query=system_number,exact,{system_number}"
 
            found,book_info = api_response(url)
            
            if found:
                print("found the book")

                write_row(book_info, sheet,row_num, batch_name, title)
            else:
                print("failed to find the book")



# excel_completer(r"excel/all years.xlsx")