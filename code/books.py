import re
from window import SentenceComparer
from docx import Document
import openpyxl
import requests
import xlsxwriter
import os

from fuzzywuzzy import fuzz



def get_books_names(doc_file_path):
    doc = Document(doc_file_path)
    titles = []
    books_title = []
    books_in_sheet = []
    authors = []
    books = 0
    years = []
    count_books = 0
    check = 0

    count = 0
    book_title = ""

    for paragraph in doc.paragraphs:
        count = 0
        book_title = ""
        all_bold = True
        text = paragraph.text

        for run in paragraph.runs:

            if run.font.underline:
                print("underline")
                pass

            elif run.bold:
                book_title = book_title + run.text

            else:
                all_bold = False

        if (
            all_bold and paragraph.text.strip() and ">>" not in paragraph.text
        ):  # checks if the line is a title or a book name
            title = paragraph.text
            if check == 1 and books == 0:
                print("pop")
                titles.pop()
                titles.append(title)
                check = 0

            elif books == 0:
                check += 1
                titles.append(title)

            if books != 0 and ">>" not in title:
                titles.append(title)
                books_in_sheet.append(books)

                books = 0
                check = 0

        elif book_title:
            book_title = book_title.strip(" ")
            book_title = book_title.strip(",")
            books_title.append(book_title)
            year = re.findall(r"\d+", paragraph.text)
            if year == []:
                years.append(0)
            elif isinstance(year, list):
                years.append(year[0])
            else:
                years.append(year)
            books += 1
            count_books += 1
            check = 0

            author = get_author(text, book_title)

            authors.append(author)

    if books > 0:
        books_in_sheet.append(books)

    return books_title, titles, books_in_sheet, authors, years


def get_author(paragraph, book_name):
    paragraph = paragraph.replace(book_name, "")
    words = paragraph.split(",")

    words = [word.strip(" ") for word in words]

    for word in words:
        if word != "" and ("2" not in word and "1" not in word):
            return word


def book_name_normelized(book_title):

    new_book_title = ""
    for letter in book_title:
        if ord(letter) < 1400:
            letter = hex(ord(letter)).replace("0x", "%")
            new_book_title = new_book_title + letter

        else:
            new_book_title = new_book_title + letter

    return new_book_title

def get_hebrew_date(date):
    year = 1240 
    date = date.replace('"', "")
    date = date.replace('\\', "")
    
    for letter in date:
        year = year + HEBREW_GEMATRIA[letter]

    return year
    
            
    
def get_date(date):
    year = re.findall(r"\d+", date)
    
    if year:
        year = year[0]
        return year

    text = date.split(" ")
    
    for year in text:
        if '"' in year:
           return get_hebrew_date((str)(year))

            

    return False


def find_book_by_year(found_date=0, year=0):
    found_date = (str)(get_date(found_date))
    year = (str)(year)
    
    if not found_date:
        return False

    if year != 0 and found_date != 0:
        if year not in found_date:
            return False

    return True

def find_book_by_author(author,book_title, found_title):

    book_name_found = found_title.split("/")[0]

    new_author = author.replace("-", " ")
    new_author = new_author.split(" ")
    # print(new_author)
    # print(author)
    for name in new_author:
        if name not in found_title:
            return False

    if book_title in book_name_found:
        return True
    return True

def find_book(author, found_title , found_date = 0, year = 0):
    if find_book_by_year(found_date, year) and find_book_by_author(author, found_title):
        return True
    
    if find_book_by_year(found_date, year):
        return True

    return False
    

    

def api_response(base_url, book_title, author, year):
    response = requests.get(base_url)
    new_books_data = []
    error_mes = []
    if response.status_code == 200 and response.json() != [] and year != 0:
        books_data = response.json()
        
        if len(books_data) == 1:
            book_data = data_organizer(books_data)
            return True, book_data

        
        if len(books_data) > 1:
            
            for book in books_data:
                book_data = data_organizer([book])
                try:
                    # found_book = find_book(author, book_data["title"], book_data["date"], year)
                    found_book = find_book_by_year(book_data["date"], year)
                except Exception as e:
                    found_book = False

                if found_book:
                    new_books_data.append(book_data)


            if len(new_books_data) == 0:
                
                for book in books_data:
                    book_data = data_organizer([book])
                    found_book = find_book_by_author(author,book_title, book_data["title"])

                    if found_book:
                        new_books_data.append(book_data)


                        
            if len(new_books_data) == 1:
                return True, new_books_data[0]


            if len(new_books_data) > 1:
                print("found 2 books with the same date")

                for book_data in new_books_data:
                
                        found_book = find_book_by_author(author,book_title, book_data["title"])
                        
                        if not found_book:
                            print("found a book with the same date but wrong author")
                            new_books_data.remove(book_data)

            if len(new_books_data) == 1 or len(new_books_data) > 1:
                return True, new_books_data[0] 


    return False, None


def search_book_in_database(book_title, author, year):
    changed_book_title = book_name_normelized(book_title)
    changed_author = book_name_normelized(author)

    if year != 0:
        first_url = f"https://api.nli.org.il/openlibrary/search?api_key={USER_KEY}&query=title,exact,{changed_book_title},AND;language,contains,heb&material_type=book"
        second_url = f"https://api.nli.org.il/openlibrary/search?api_key={USER_KEY}&query=title,contains,{changed_book_title},AND;language,contains,heb&material_type=book"
        third_url = f"https://api.nli.org.il/openlibrary/search?api_key={USER_KEY}&query=title,contains,{changed_book_title},AND;language,contains,heb,AND;creator,contains,{changed_author},AND;start_date,contains,{year}&material_type=book"
        urls = [third_url, first_url, second_url]

        for url in urls:
            # print(f"{year} ---  {book_title}")
            response, data = api_response(url, book_title, author, year)
            if response:
                return data

            # print(url)

    return False


def data_organizer(book_data):
    if not book_data or not isinstance(book_data, list) or len(book_data) == 0:
        return {}

    data = book_data[0]
    book_info = {}

    for detail, values in data.items():
        field = detail.replace("http://purl.org/dc/elements/1.1/", "")

        if field == "non_standard_date":
            field = "date"

        if field not in FIELDS:
            continue

        for value_dict in values:
            value = value_dict.get("@value") or value_dict.get("@id")

            if value:
                if field in book_info:
                    if isinstance(book_info[field], list):
                        if value not in book_info[field]:
                            book_info[field].append(value)
                    else:
                        book_info[field] = [book_info[field], value]
                else:
                    book_info[field] = value

    return book_info


def clean_sheet_title(title):

    title = title.replace("*", "")
    title = title.replace("הספרים המוצעים", "")
    title = title.replace("'", "")

    return title


def delete_file_if_exists(filepath):
    # Check if the file exists
    if os.path.exists(filepath):
        try:
            # Attempt to delete the file
            os.remove(filepath)
            print(f"File '{filepath}' has been deleted.")
        except OSError as e:
            print(f"Error deleting file '{filepath}': {e}")
    else:
        print(f"File '{filepath}' does not exist.")


def get_xlsx(doc_file_path):
    books_title, titles, books_in_sheet, authors, years = get_books_names(doc_file_path)

    count = 0
    xlsx_file_path = r"excel/" + os.path.basename(doc_file_path).replace(
        ".docx", ".xlsx"
    )
    text_file_path = r"text/" + os.path.basename(doc_file_path).replace(".docx", ".txt")

    delete_file_if_exists(xlsx_file_path)
    delete_file_if_exists(text_file_path)

    workbook = xlsxwriter.Workbook(xlsx_file_path)
    workbook.close()
    workbook = openpyxl.load_workbook(xlsx_file_path)

    # Remove the default "Sheet" if it's still there
    if "Sheet1" in workbook.sheetnames:
        workbook.remove(workbook["Sheet1"])

    with open(text_file_path, "w", encoding="utf-8") as file:
        for num, title in enumerate(titles):

            sheet_title = clean_sheet_title(title)

            sheet = workbook.create_sheet(sheet_title)

            # sheet = workbook.active
            print(f"created a new sheet: {sheet_title}")
            row_num = 2

            # Write the headers to the first row of the sheet
            for col_num, field in enumerate(FIELDS, start=1):
                sheet.cell(row=1, column=col_num).value = field

            for temp in range(books_in_sheet[num]):  # Example: 20 books

                print(f"{count + 1} / {len(books_title)}")

                book_title = books_title[count]
                author = authors[count]
                year = years[count]

                book_info = search_book_in_database(book_title, author, year)
                if book_info:
                    book = []

                    for field in FIELDS:
                        if field in book_info:
                            if isinstance(book_info[field], list):
                                x = True
                                for item in book_info[field]:
                                    if item is not None and x:
                                        book_info[field] = item
                                        x = False

                            book.append(book_info[field])
                        else:
                            book.append(" ")

                    for col_num, item in enumerate(book, start=1):

                        sheet.cell(row=row_num, column=col_num).value = item

                    row_num += 1
                    count += 1

                else:
                    file.write(f"{book_title} / {author}  {year}\n")
                    print(f"didnt find: {book_title} / {author}")
                    sheet.cell(row=row_num, column=1).value = f"{book_title} / {author}"
                    row_num += 1
                    count += 1

            workbook.save(xlsx_file_path)


# USER_KEY = "8rqp3wBL0YX2yCaNPOSha2bhbrJ2Zd2Fagw8vuGw"
USER_KEY = "g5ay94wdXhFvZQgpHbsLyaEHW3iKi2a90vG2eUI1"
# USER_KEY = "DVQyidFLOAjp12ib92pNJPmflmB5IessOq1CJQDK"

HEBREW_GEMATRIA = {
    'א': 1,
    'ב': 2,
    'ג': 3,
    'ד': 4,
    'ה': 5,
    'ו': 6,
    'ז': 7,
    'ח': 8,
    'ט': 9,
    'י': 10,
    'כ': 20,
    'ל': 30,
    'מ': 40,
    'נ': 50,
    'ס': 60,
    'ע': 70,
    'פ': 80,
    'צ': 90,
    'ק': 100,
    'ר': 200,
    'ש': 300,
    'ת': 400,
    'ך': 20,  # Final kaf
    'ם': 40,  # Final mem
    'ן': 50,  # Final nun
    'ף': 80,  # Final pe
    'ץ': 90   # Final tzadi
}

FIELDS = [
    "title",
    "contributor",
    "identifier",
    "linkToMarc",
    "creator",
    "subject",
    "thumbnail",
    "format",
    "date",
    "publisher",
    "language",
    "recordid",
    "type",
    "source",
]

if __name__ == "__main__":

    path_to_docx_dir = r"docs\completed"
    docx_dir_list = os.listdir(path_to_docx_dir)

    for docs in docx_dir_list:
        doc_file_path = r"docs\completed\\" + docs
        print(docs)
        get_xlsx(doc_file_path)

    # doc_file_path = "docs/2015.docx"
    # books_title, titles, books_in_sheet, authors, years = get_books_names(doc_file_path)

    # print(f"{books_title[49]} {years[49]}")
    # author = "שם-טוב תמי"
    # text = ""
    # new_author = author.replace("-", " ")
    # new_author = new_author.split(" ")

    # for name in new_author:
    #     if name not in text:
    #         print(name)
    #         break

    # year = 'שדגג תשע"ח' 

    # print(get_date(year))
        
    path_to_text_dir = r"docs"
    text_dir_list = os.listdir(path_to_text_dir)

    # print(author_check("שם-טוב תמי", "מארי קירי שגלתה את הקרניים החזקות בעולם / כתבה תמי שם-טוב"))


# get_base_url(book_title,author)
# book_title = "אבא, אמא"
# print(book_name_normelized(book_title))
