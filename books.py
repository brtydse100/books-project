import csv
from plistlib import InvalidFileException
import re
import time
import zipfile
from docx import Document
from docx.shared import Pt  
import openpyxl
import requests
from openpyxl import load_workbook
import xlsxwriter
import os



def get_books_names(doc_file_path):
    doc = Document(doc_file_path)
    titles = []
    books_title = []
    books_in_sheet = []
    authors = []
    books = 0
    count_books = 0
    check = 0

    count = 0
    book_title = ""

    for paragraph in doc.paragraphs:
        count = 0
        book_title = ""
        all_bold = True
        text = paragraph.text

        for run in paragraph.runs:

            
            if run.font.underline:
                print("underline")
                pass
            
            elif run.bold:
                book_title = book_title + run.text
                
            else:
                all_bold = False
            


        if all_bold and paragraph.text.strip() and ">>" not in paragraph.text: # checks if the line is a title or a book name
            title = paragraph.text
            if check == 1 and books == 0:
                print("pop")
                titles.pop()
                titles.append(title)
                check = 0
                
            elif books == 0:
                check +=1
                titles.append(title)
                
            if(books != 0 and ">>" not in title):
                titles.append(title)
                books_in_sheet.append(books)
 
                books = 0
                check = 0

        
        elif book_title:  

            books_title.append(book_title)
            books +=1
            count_books +=1
            check = 0
            
            author = get_author(text,book_title)
            
            authors.append(author)
            

        
    if(books > 0):
        books_in_sheet.append(books)

    print(books_in_sheet)
    
    return books_title, titles, books_in_sheet ,authors

def get_author(paragraph, book_name):
    # print(paragraph)
    paragraph = paragraph.replace(book_name, "")
    words = paragraph.split(",")
    
    words = [word.strip(' ') for word in words]
    
    for word in words:
        if word != "" and ("2" not in word and "1" not in word):
            return word

def book_name_normelized(book_title):
    book_title = book_title.strip(" ")
    book_title = book_title.strip(",")
    print(book_title)
    new_book_title = ""
    for letter in book_title:
        if (ord(letter) < 1400):
            letter = hex(ord(letter)).replace("0x", "%")
            new_book_title = new_book_title + letter

        else:
            new_book_title = new_book_title + letter

    new_book_title = "title,exact," + new_book_title
    
    return new_book_title
    
def search_book_in_database(book_title):
    book_title = book_name_normelized(book_title)
    
    base_url = f"https://api.nli.org.il/openlibrary/search?api_key={USER_KEY}&query={book_title}"
    field_list = []
    book_info = {}
    response = requests.get(base_url)
    # print(base_url)
    if response.status_code != 200:
        return book_title
    book_data = response.json()
    
    for data in book_data:
        for detail in data:
            field = detail.replace("http://purl.org/dc/elements/1.1/", "")
            field_list = []
                
            if(type(data[detail][0]) is dict):
                
                try:
                    value = data[detail][0]["@value"]
                        
                except Exception as err :
                    value = data[detail][0]["@id"]


                if field in book_info:
                    if(type(book_info[field]) is not list):
                        
                        field_list.append(book_info[field])
                        field_list.append(value)
                        book_info[field] = field_list
                        
                    elif value not in book_info[field]:
                        book_info[field] = book_info[field].append(value)

                elif field in FIELDS:
                    book_info[field] = value

    return book_info

def clean_sheet_title(title):
        
        title = title.replace("*", "")
        title = title.replace("הספרים המוצעים", "")
        title = title.replace("'", "")

        return title

def check_book_title(book_title, author):
    words = book_title.split()  # Split the phrase into words
    if len(words) > 2:
        for i in range(len(words)):
            new_book_title = " ".join(words[i:])  # Join words starting from index i to the end
            book_info = search_book_in_database(new_book_title)
            
            if (book_info and not isinstance(book_info,str)):
                if(compare_title(author, book_info["title"])):
                    return book_info
            
    return False

def delete_file_if_exists(filepath):
    # Check if the file exists
    if os.path.exists(filepath):
        try:
            # Attempt to delete the file
            os.remove(filepath)
            print(f"File '{filepath}' has been deleted.")
        except OSError as e:
            print(f"Error deleting file '{filepath}': {e}")
    else:
        print(f"File '{filepath}' does not exist.")

def compare_title(author, found_title):
    
    if found_title is None:
        return False

    
    if(isinstance(found_title, list)):
        found_title = " ".join([str(item) if item is not None else "" for item in found_title])
        

    found_title = set(found_title)
    author = set(author)

    common_words = found_title.intersection(author)

    if(len(common_words) <= len(author) + 1):
        return True

    return False
    

def get_xlsx(doc_file_path):
    books_title, titles, books_in_sheet,authors = get_books_names(doc_file_path)
    count = 0
    # print(titles)
    xlsx_file_path = r"excel/" + os.path.basename(doc_file_path).replace(".docx", ".xlsx")

    delete_file_if_exists(xlsx_file_path) 

    workbook = xlsxwriter.Workbook(xlsx_file_path)
    workbook.close()
    workbook = openpyxl.load_workbook(xlsx_file_path)

    # Remove the default "Sheet" if it's still there
    if "Sheet1" in workbook.sheetnames:
        workbook.remove(workbook["Sheet1"])

        
    for num, title in enumerate(titles):
        
        sheet_title = clean_sheet_title(title)
        
        sheet = workbook.create_sheet(sheet_title)
        
        # sheet = workbook.active
        print(f"created a new sheet: {sheet_title}")
        row_num = 2
        
        # Write the headers to the first row of the sheet
        for col_num, field in enumerate(FIELDS, start=1):
            sheet.cell(row=1, column=col_num).value = field

        
        for temp in range(books_in_sheet[num]):  # Example: 20 books
            
            print(f"{count + 1} / {len(books_title)}")
            book_title = books_title[count]
            author = authors[count]

            
            book_info = search_book_in_database(book_title)
            if (book_info):
                book = []
                for field in FIELDS:
                    try:
                        if field == "title":
                            
                            if(book_info[field] is None or book_info[field] == ""):
                                print("no title adding one")
                                print(book_title)
                                
                                book.append(book_title)
                            else:
                                book.append(book_info[field])
                        else:
                            book.append(book_info[field])
                    except Exception:
                        pass
                for col_num, item in enumerate(book, start=1):
                    if isinstance(item, list):
                        item = set(item)
                        item = list(item)
                        
                        values = ','.join(str(v) for v in item if v is not None)
                        
                        sheet.cell(row=row_num, column=col_num).value = values
                    else:
                        sheet.cell(row=row_num, column=col_num).value = item

                row_num += 1
                count += 1
                
            else:
                
                book_info = check_book_title(book_title, author)

                if book_info:

                    print(f"managed to retrieve the data for: {book_title} under the title: {book_info["title"]}")
                    book = []
                    
                    print(book_info["title"])
                    
                    for field in FIELDS:
                        try:
                            if field == "title":
                                
                                if(book_info[field] is None or book_info[field] == ""):
                                    print("no title adding one")
                                    print(book_title)
                                    
                                    
                                    book.append(book_title)
                                else:
                                    book.append(book_info[field])
                            else:
                                book.append(book_info[field])
                        except Exception:
                            pass
                    for col_num, item in enumerate(book, start=1):
                        if isinstance(item, list):
                            item = set(item)
                            item = list(item)
                            
                            values = ','.join(str(v) for v in item if v is not None)
                            
                            sheet.cell(row=row_num, column=col_num).value = values
                        else:
                            sheet.cell(row=row_num, column=col_num).value = item

                    row_num += 1
                    count += 1


                else:
                    print(f"failed to find the book: {book_title}")
                    print("adding the title")
                    print(book_title)
                    
                    sheet.cell(row=row_num, column=1).value = book_title
                    row_num += 1
                    count += 1

        workbook.save(xlsx_file_path)



USER_KEY = "8rqp3wBL0YX2yCaNPOSha2bhbrJ2Zd2Fagw8vuGw"
USER_KEY = "DVQyidFLOAjp12ib92pNJPmflmB5IessOq1CJQDK"
FIELDS = ['title', 'contributor', 'identifier', 'linkToMarc', 'creator', 'subject',  'thumbnail', 'format', 'date', 'publisher', 'language', 'recordid', 'type', 'source']
# https://api.nli.org.il/openlibrary/search?api_key=8rqp3wBL0YX2yCaNPOSha2bhbrJ2Zd2Fagw8vuGw&query=title,exact,החתול%20שרצה%20להיות%20איש%2c%20
book_title = "התינוקת מסלובניה"
# book_title = "הפרעוש: סיפורו המדהים של ליאו מסי"
book_title = book_title.replace(" ", "%20")
book_title = "title,exact,"+ book_title

# book_info = search_book_in_database(book_title)
# print(book_info)
path = r"docs"
dir_list = os.listdir(path)

# for docs in dir_list:
#     doc_file_path = r"docs\\" + docs
#     print(docs)
#     get_xlsx(doc_file_path)

# doc_file_path = r"docs\done\2015.docx"
doc_file_path = r"C:\Users\Ido\Desktop\books-project\docs\done\booklist2019.docx"

get_xlsx(doc_file_path)
# books_title, titles, books_in_sheet = get_books_names(doc_file_path)
# print(books_title)

title1 = [None, 'פינות אפלות / איי. אם. מאדן ; תרגום: ליאורה כרמלי ; עריכה: לין תהל כהן.']
author = "יובל-יאיר שירלי"


# print(compare_title(author, title1))
